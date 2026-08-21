"""Exportação de relatórios executivos FieldNode em formato Excel."""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from api_tcc.models import LeituraTelemetria, RegistroAnalise


class RelatorioExportarView(APIView):
    """Gera um relatório .xlsx com resumo, telemetria e eventos de IA."""

    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    verde_fieldnode = "1A3326"
    cinza_fieldnode = "2D3748"

    def get(self, request):
        maquina_id = request.query_params.get("maquina_id")
        if not maquina_id:
            return HttpResponse('Parâmetro "maquina_id" é obrigatório', status=400)

        leituras = LeituraTelemetria.objects.filter(maquina_id=maquina_id).order_by(
            "timestamp"
        )
        analises = RegistroAnalise.objects.filter(maquina_id=maquina_id).order_by(
            "criado_em"
        )

        workbook = Workbook()
        resumo = workbook.active
        resumo.title = "Resumo"
        self._montar_aba_resumo(resumo, maquina_id, leituras, analises)

        telemetria = workbook.create_sheet("Telemetria")
        self._montar_aba_dados(telemetria, leituras)

        eventos = workbook.create_sheet("Eventos")
        self._montar_aba_eventos(eventos, analises)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type=self.content_type)
        response["Content-Disposition"] = (
            f'attachment; filename="fieldnode_relatorio_{maquina_id}.xlsx"'
        )
        return response

    def _montar_aba_resumo(self, ws, maquina_id, leituras, analises):
        ws["A1"] = "FieldNode — Relatório Executivo de Telemetria"
        ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=self.verde_fieldnode)
        ws["A1"].alignment = Alignment(vertical="center")
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 28

        metricas = (
            ("Identificador da Máquina", maquina_id),
            ("Total de Leituras Capturadas", leituras.count()),
            ("Alertas Críticos Registrados", analises.filter(status="CRITICO").count()),
            ("Alertas de Atenção Registrados", analises.filter(status="ATENCAO").count()),
        )
        for row, (titulo, valor) in enumerate(metricas, start=3):
            ws.cell(row=row, column=1, value=titulo).font = Font(name="Calibri", bold=True)
            ws.cell(row=row, column=2, value=valor)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")

        for col, largura in zip("ABCD", (30, 22, 22, 22)):
            ws.column_dimensions[col].width = largura

    def _aplicar_cabecalho(self, ws, cabecalhos, cor):
        preenchimento = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        fonte = Font(name="Calibri", color="FFFFFF", bold=True)
        for coluna, titulo in enumerate(cabecalhos, start=1):
            celula = ws.cell(row=1, column=coluna, value=titulo)
            celula.fill = preenchimento
            celula.font = fonte
            celula.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _montar_aba_dados(self, ws, leituras):
        cabecalhos = ("Timestamp", "Temperatura (°C)", "Vibração", "RPM")
        self._aplicar_cabecalho(ws, cabecalhos, self.verde_fieldnode)

        if not leituras.exists():
            ws.cell(row=2, column=1, value="Nenhuma leitura registrada para esta máquina.")
        else:
            for linha, leitura in enumerate(leituras.iterator(), start=2):
                timestamp = (
                    leitura.timestamp.strftime("%d/%m/%Y %H:%M:%S")
                    if leitura.timestamp
                    else ""
                )
                ws.cell(row=linha, column=1, value=timestamp)
                ws.cell(row=linha, column=2, value=leitura.temperatura)
                ws.cell(row=linha, column=3, value=leitura.vibracao)
                ws.cell(row=linha, column=4, value=leitura.rpm)

            # Destaca temperaturas críticas no Excel e em leitores compatíveis.
            ws.conditional_formatting.add(
                f"B2:B{ws.max_row}",
                CellIsRule(
                    operator="greaterThan",
                    formula=["85"],
                    fill=PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
                    font=Font(color="9B1C1C", bold=True),
                ),
            )
            ws.auto_filter.ref = f"A1:D{ws.max_row}"

        self._definir_larguras(ws, (22, 18, 14, 12))

    def _montar_aba_eventos(self, ws, analises):
        cabecalhos = (
            "Data/Hora",
            "Status do Risco",
            "Motivos da Anomalia",
            "Recomendação Operacional",
        )
        self._aplicar_cabecalho(ws, cabecalhos, self.cinza_fieldnode)

        if not analises.exists():
            ws.cell(row=2, column=1, value="Nenhum evento anômalo registrado no período.")
        else:
            for linha, analise in enumerate(analises.iterator(), start=2):
                timestamp = (
                    analise.criado_em.strftime("%d/%m/%Y %H:%M:%S")
                    if analise.criado_em
                    else ""
                )
                ws.cell(row=linha, column=1, value=timestamp)
                ws.cell(row=linha, column=2, value=analise.status)
                motivos = analise.motivos if isinstance(analise.motivos, list) else []
                ws.cell(row=linha, column=3, value="; ".join(map(str, motivos)))
                ws.cell(row=linha, column=4, value=analise.recomendacao or "")
                for coluna in (3, 4):
                    ws.cell(row=linha, column=coluna).alignment = Alignment(wrap_text=True, vertical="top")
            ws.auto_filter.ref = f"A1:D{ws.max_row}"

        self._definir_larguras(ws, (22, 16, 45, 45))

    @staticmethod
    def _definir_larguras(ws, larguras):
        for coluna, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(coluna)].width = largura
