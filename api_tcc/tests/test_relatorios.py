"""Testes de borda para a exportação XLSX do relatório executivo."""

from datetime import datetime
from io import BytesIO
from time import monotonic

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status

from api_tcc.models import (
    AlturadoCorte,
    Colheitadeira,
    Combustivel,
    EstadodeMovimento,
    LeituraTelemetria,
    Marca,
    Modelo,
    Operario,
    PressaoPneus,
    PressaodoCorte,
    RegistroAnalise,
    StatusdeOperacao,
    TemperaturaMaquina,
    TempUmi_Ambiente,
    UnidadedeMedida,
)


class RelatorioExportarBordaTestCase(TestCase):
    """O endpoint deve sempre entregar um XLSX legível ou um erro explícito."""

    @classmethod
    def setUpTestData(cls):
        unidade = UnidadedeMedida.objects.create(nome="Centímetro")
        marca = Marca.objects.create(nome="Marca de teste")
        cls.modelo = Modelo.objects.create(nome="Modelo de teste", marca=marca)
        cls.combustivel = Combustivel.objects.create(tipo="Diesel", porcentagem=100)
        cls.pressao_pneus = PressaoPneus.objects.create(
            pressao=2.5, unidade_de_medida=unidade
        )
        cls.altura_corte = AlturadoCorte.objects.create(
            altura=5, unidade_de_medida=unidade
        )
        cls.pressao_corte = PressaodoCorte.objects.create(
            pressao=30, unidade_de_medida=unidade
        )
        cls.temp_ambiente = TempUmi_Ambiente.objects.create(temperatura=25, umidade=60)
        cls.temp_maquina = TemperaturaMaquina.objects.create(
            temperatura=80, maquina=cls.modelo
        )
        cls.operario = Operario.objects.create(
            nome="Operário de teste", tempo_de_servico=1, no_banco=True
        )
        cls.status_operacao = StatusdeOperacao.objects.create(
            em_operacao=True, tempo_de_operacao=1
        )
        cls.estado_movimento = EstadodeMovimento.objects.create(
            em_movimento=False, velocidade=0
        )

    def setUp(self):
        self.url = reverse("relatorio-exportar")

    def criar_maquina(self, maquina_id):
        return Colheitadeira.objects.create(
            modelo=self.modelo,
            maquina_id=maquina_id,
            combustivel=self.combustivel,
            pressao_pneus=self.pressao_pneus,
            altura_do_corte=self.altura_corte,
            pressao_do_corte=self.pressao_corte,
            temp_umi_ambiente=self.temp_ambiente,
            temperatura_maquina=self.temp_maquina,
            operario=self.operario,
            status_de_operacao=self.status_operacao,
            estado_de_movimento=self.estado_movimento,
        )

    def obter_planilha(self, response):
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return load_workbook(BytesIO(response.content))

    def test_maquina_sem_parametro_retorna_400(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("maquina_id", response.content.decode())

    def test_maquina_inexistente_retorna_404_claro(self):
        response = self.client.get(self.url, {"maquina_id": "inexistente-01"})
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn("Máquina não encontrada", response.content.decode())

    def test_maquina_sem_leituras_gera_xlsx_com_resumo_zero(self):
        maquina_id = "maquina-vazia-01"
        self.criar_maquina(maquina_id)

        response = self.client.get(self.url, {"maquina_id": maquina_id})
        self.assertIn(
            f"fieldnode_relatorio_{maquina_id}.xlsx", response["Content-Disposition"]
        )
        workbook = self.obter_planilha(response)

        self.assertEqual(workbook.sheetnames, ["Resumo", "Telemetria", "Eventos"])
        self.assertEqual(workbook["Resumo"]["B4"].value, 0)
        self.assertEqual(
            workbook["Telemetria"]["A2"].value,
            "Nenhuma leitura registrada para esta máquina.",
        )
        self.assertEqual(workbook["Telemetria"]["A1"].fill.fgColor.rgb, "001A3326")
        self.assertEqual(workbook["Eventos"]["A1"].fill.fgColor.rgb, "002D3748")

    def test_maquina_com_uma_leitura_e_evento_gera_xlsx_compativel(self):
        maquina_id = "maquina-unitaria-01"
        self.criar_maquina(maquina_id)
        instante = timezone.make_aware(datetime(2026, 8, 21, 10, 30))
        LeituraTelemetria.objects.create(
            maquina_id=maquina_id,
            temperatura=70.0,
            vibracao=1.2,
            rpm=1500,
            timestamp=instante,
        )
        RegistroAnalise.objects.create(
            maquina_id=maquina_id,
            status="NORMAL",
            motivos=[],
            recomendacao="Operação normal.",
        )

        workbook = self.obter_planilha(
            self.client.get(self.url, {"maquina_id": maquina_id})
        )
        self.assertEqual(workbook["Resumo"]["B4"].value, 1)
        self.assertIsInstance(workbook["Telemetria"]["A2"].value, datetime)
        self.assertEqual(
            workbook["Telemetria"]["A2"].number_format, "DD/MM/YYYY HH:MM:SS"
        )
        self.assertEqual(workbook["Telemetria"]["B2"].value, 70.0)
        self.assertEqual(workbook["Eventos"]["B2"].value, "NORMAL")

    def test_performance_com_cinco_mil_leituras(self):
        maquina_id = "maquina-bulk-01"
        self.criar_maquina(maquina_id)
        instante = timezone.now()
        LeituraTelemetria.objects.bulk_create(
            [
                LeituraTelemetria(
                    maquina_id=maquina_id,
                    temperatura=75.0 + (indice % 15),
                    vibracao=2.0,
                    rpm=1800,
                    timestamp=instante,
                )
                for indice in range(5000)
            ],
            batch_size=1000,
        )

        inicio = monotonic()
        response = self.client.get(self.url, {"maquina_id": maquina_id})
        duracao = monotonic() - inicio

        workbook = self.obter_planilha(response)
        self.assertEqual(workbook["Resumo"]["B4"].value, 5000)
        self.assertEqual(workbook["Telemetria"].max_row, 5001)
        self.assertLess(duracao, 4.0, f"Demorou {duracao:.2f}s para gerar o relatório em massa.")
